"""
CSV/Excel Import Service
Ports V1 data loading logic to DB-based import with upsert support.
pandas/openpyxl imported lazily inside functions to reduce startup memory.
"""
from __future__ import annotations
import io
import logging
from datetime import datetime
from sqlalchemy.orm import Session
from sqlalchemy import and_

logger = logging.getLogger("rodmat.import_service")

from app.models.sales import SalesOrder, AffiliateSale
from app.models.product import Product
from app.models.combo import Combo, ComboItem
from app.models.inventory import InitialInventory, IncomingStock



# ─── BRAND resolution helper (multi-brand LuxPerfumes support) ───────────────
_BRAND_MAP_CACHE: dict[str, tuple[dict, dict]] = {}  # store_id → (sku_to_brand, prefix_to_brand)

def _get_brand_maps(db, store_id: str):
    """Return (sku_to_brand_id, prefix_to_brand_id) for the store.
    Cached per-call to avoid N queries during bulk import.
    Empty dicts if brands_enabled=False for the store."""
    if store_id in _BRAND_MAP_CACHE:
        return _BRAND_MAP_CACHE[store_id]
    from sqlalchemy import text as _t
    # SKU → brand_id map (source of truth via products.brand_id)
    rows = db.execute(_t("""
        SELECT p.sku, p.brand_id
        FROM products p
        WHERE p.store_id = :sid AND p.brand_id IS NOT NULL
    """), {"sid": store_id}).fetchall()
    sku_map = {r[0]: r[1] for r in rows if r[0]}
    # Prefix fallback: build from brands.sku_prefixes_note (comma-separated prefixes like 'AV-*, LAT-*')
    prefix_map: dict[str, str] = {}
    brand_rows = db.execute(_t("""
        SELECT id, sku_prefixes_note FROM brands WHERE store_id = :sid AND is_active = TRUE
    """), {"sid": store_id}).fetchall()
    for bid, note in brand_rows:
        if not note: continue
        # Extract prefixes like "AV", "AT", "LAT" from "AV-*" or "AT-*, LAT-*"
        import re as _re
        for prefix in _re.findall(r"([A-Z]{2,5})-\*", note.upper()):
            prefix_map[prefix] = bid
    _BRAND_MAP_CACHE[store_id] = (sku_map, prefix_map)
    return sku_map, prefix_map


def _resolve_brand_id(sku: str | None, sku_map: dict, prefix_map: dict) -> str | None:
    """Resolve brand_id for a SKU:
    1) Exact SKU match in products → use products.brand_id (source of truth)
    2) Prefix match (AV-*, AT-*, LAT-*) → fallback for combos/variants no dados de alta
    3) None → row stays NULL (log-worthy)"""
    if not sku:
        return None
    if sku in sku_map:
        return sku_map[sku]
    # Prefix fallback: extraer "AV" de "AV-RG" o "AV-RPG/2"
    import re as _re
    m = _re.match(r"^([A-Z]{2,5})-", sku.upper())
    if m and m.group(1) in prefix_map:
        return prefix_map[m.group(1)]
    return None

def _detect_separator(content: bytes) -> str:
    first_line = content.split(b'\n')[0].decode('utf-8-sig', errors='replace')
    return '\t' if first_line.count('\t') > first_line.count(',') else ','


def _safe_float(val) -> float | None:
    try:
        v = float(val)
        return v if v == v else None  # NaN check without pandas
    except (ValueError, TypeError):
        return None


def _safe_int(val) -> int:
    try:
        v = float(val)
        return int(v) if v == v else 0
    except (ValueError, TypeError):
        return 0


def _safe_str(val) -> str | None:
    if val is None:
        return None
    try:
        import math
        if isinstance(val, float) and math.isnan(val):
            return None
    except Exception:
        pass
    s = str(val).strip()
    return s if s and s != 'nan' else None


def _safe_datetime(val, dayfirst: bool = False) -> datetime | None:
    try:
        import pandas as pd
        dt = pd.to_datetime(val, errors='coerce', dayfirst=dayfirst)
        return dt.to_pydatetime() if pd.notna(dt) else None
    except Exception:
        return None


def parse_orders_csv(content: bytes, store_id: str, db: Session, batch_id: str | None = None) -> dict:
    """Parse TikTok orders CSV (AllBBDD format). Replaces TikTok orders only (keeps Amazon).
    Multi-brand aware: resuelve brand_id por SKU (product.brand_id) o fallback por prefijo."""
    import uuid as _uuid
    import pandas as pd
    from sqlalchemy.dialects.postgresql import insert as pg_insert

    # Brand map (fresh per import — invalidate cache in case products/brands cambiaron)
    _BRAND_MAP_CACHE.pop(store_id, None)
    _sku_map, _prefix_map = _get_brand_maps(db, store_id)

    sep = _detect_separator(content)
    df = pd.read_csv(
        io.BytesIO(content), sep=sep,
        encoding='utf-8-sig', on_bad_lines='skip', engine='python'
    )
    df.columns = df.columns.str.strip()

    rows = []
    errors = 0

    for _, row in df.iterrows():
        try:
            order_id = _safe_str(row.get('Order ID'))
            if not order_id:
                errors += 1
                continue

            sku_id = _safe_str(row.get('SKU ID'))
            created_time = _safe_datetime(row.get('Created Time'))
            shipped_time = _safe_datetime(row.get('Shipped Time'))

            rows.append(dict(
                id=str(_uuid.uuid4()),
                store_id=store_id,
                tiktok_order_id=order_id,
                order_date=created_time,
                sku=sku_id,
                seller_sku=_safe_str(row.get('Seller SKU')),
                product_name=_safe_str(row.get('Product Name')),
                quantity=_safe_int(row.get('Quantity', 1)),
                status=_safe_str(row.get('Order Status')),
                substatus=_safe_str(row.get('Order Substatus')),
                price=_safe_float(row.get('SKU Unit Original Price')),
                shipped_time=shipped_time,
                created_time=created_time,
                sku_subtotal_after_discount=_safe_float(row.get('SKU Subtotal After Discount')),
                order_amount=_safe_float(row.get('Order Amount')),
                order_refund_amount=_safe_float(row.get('Order Refund Amount')),
                shipping_fee_after_discount=_safe_float(row.get('Shipping Fee After Discount')),
                original_shipping_fee=_safe_float(row.get('Original Shipping Fee')),
                sku_seller_discount=_safe_float(row.get('SKU Seller Discount')),
                sku_platform_discount=_safe_float(row.get('SKU Platform Discount')),
                cancelation_return_type=_safe_str(row.get('Cancelation/Return Type')),
                fulfillment_type=_safe_str(row.get('Fulfillment Type')),
                buyer_username=_safe_str(row.get('Buyer Username')),
                variation=_safe_str(row.get('Variation')),
                recipient=_safe_str(row.get('Recipient')),
                city=_safe_str(row.get('City')),
                state=_safe_str(row.get('State')),
                platform='tiktok',
                import_batch_id=batch_id,
                raw_data=None,
                brand_id=_resolve_brand_id(_safe_str(row.get('Seller SKU')) or sku_id, _sku_map, _prefix_map),
            ))
        except Exception:
            errors += 1

    # UPSERT by (store_id, tiktok_order_id, sku) — never deletes historical orders.
    # Orders that already exist get their status/quantity/prices updated.
    _update_cols = [
        'product_name', 'quantity', 'status', 'substatus', 'price', 'shipped_time',
        'sku_subtotal_after_discount', 'order_amount', 'order_refund_amount',
        'shipping_fee_after_discount', 'original_shipping_fee', 'sku_seller_discount',
        'sku_platform_discount', 'cancelation_return_type', 'fulfillment_type',
        'buyer_username', 'variation', 'recipient', 'city', 'state', 'import_batch_id',
        'brand_id',
    ]

    BATCH = 500   # micro-batches: evita statement_timeout Supabase Micro
    inserted, updated = 0, 0
    for i in range(0, len(rows), BATCH):
        batch = rows[i:i + BATCH]
        stmt = pg_insert(SalesOrder).values(batch)
        stmt = stmt.on_conflict_do_update(
            constraint='uq_store_order_sku',
            set_={col: getattr(stmt.excluded, col) for col in _update_cols},
        )
        db.execute(stmt)
        db.flush()
        inserted += len(batch)

    db.commit()
    return {"total_rows": len(df), "inserted": inserted, "updated": 0, "errors": errors}


def parse_affiliate_csv(content: bytes, store_id: str, db: Session) -> dict:
    """Parse affiliate/creator CSV. Upserts by (store_id, order_id, sku) using bulk INSERT ON CONFLICT."""
    import uuid
    import pandas as pd
    from sqlalchemy.dialects.postgresql import insert as pg_insert

    df = pd.read_csv(io.BytesIO(content), encoding='utf-8-sig')
    df.columns = df.columns.str.strip()

    rows = []
    errors = 0

    for _, row in df.iterrows():
        try:
            order_id = _safe_str(row.get('Order ID'))
            if not order_id:
                errors += 1
                continue

            sku = _safe_str(row.get('SKU ID', row.get('Product SKU ID')))
            # Commission rate is exported as "15%" — strip the percent sign
            comm_rate_raw = str(row.get('Standard commission rate', '') or '').replace('%', '').strip()

            rows.append(dict(
                id=str(uuid.uuid4()),
                store_id=store_id,
                order_id=order_id,
                creator_username=_safe_str(row.get('Creator Username')),
                product_name=_safe_str(row.get('Product Name')),
                sku=sku,
                quantity=_safe_int(row.get('Quantity', 1)),
                commission=_safe_float(
                    row.get('Est. standard commission payment',
                            row.get('Actual Commission Payment'))
                ),
                content_type=_safe_str(row.get('Content Type')),
                payment_amount=_safe_float(row.get('Payment Amount')),
                order_status=_safe_str(row.get('Order Status')),
                # TikTok affiliate CSVs use DD/MM/YYYY format
                time_created=_safe_datetime(row.get('Time Created'), dayfirst=True),
                commission_rate=_safe_float(comm_rate_raw),
                est_commission_base=_safe_float(row.get('Est. Commission Base')),
                raw_data=None,
            ))
        except Exception:
            errors += 1

    if not rows:
        return {"total_rows": len(df), "inserted": 0, "updated": 0, "errors": errors}

    # Bulk upsert: INSERT ... ON CONFLICT (store_id, order_id, sku) DO UPDATE
    # No N+1 SELECT queries — one batch per 1000 rows
    update_cols = ['creator_username', 'product_name', 'quantity', 'commission',
                   'content_type', 'payment_amount', 'order_status', 'time_created',
                   'commission_rate', 'est_commission_base']
    BATCH = 1000
    total = 0
    for i in range(0, len(rows), BATCH):
        batch = rows[i:i + BATCH]
        stmt = pg_insert(AffiliateSale).values(batch)
        stmt = stmt.on_conflict_do_update(
            constraint='uq_store_affiliate_order_sku',
            set_={col: getattr(stmt.excluded, col) for col in update_cols},
        )
        db.execute(stmt)
        db.flush()
        total += len(batch)

    db.commit()
    return {"total_rows": len(df), "inserted": total, "updated": 0, "errors": errors}


def parse_products_excel(content: bytes, store_id: str, db: Session) -> dict:
    """Parse Productos individualizados.xlsx. Upserts by (store_id, sku)."""
    import pandas as pd
    df = pd.read_excel(io.BytesIO(content))
    df.columns = df.columns.str.strip()

    inserted, updated, errors = 0, 0, 0

    for _, row in df.iterrows():
        try:
            producto = _safe_str(row.get('Producto'))
            if not producto:
                errors += 1
                continue

            sku = producto.strip()
            existing = db.query(Product).filter(and_(
                Product.store_id == store_id,
                Product.sku == sku,
            )).first()

            price_cost = _safe_float(row.get('Coste'))
            price_sale = _safe_float(row.get('PRECIO', row.get('Precio')))
            units_per_box = _safe_int(row.get('UNIDADES POR CAJA', row.get('Unidades por caja', 1)))
            category = _safe_str(row.get('Tipo'))
            supplier = _safe_str(row.get('Proveedor'))

            if existing:
                existing.name = sku
                existing.price_cost = price_cost
                existing.price_sale = price_sale
                existing.units_per_box = units_per_box if units_per_box else existing.units_per_box
                existing.category = category or existing.category
                existing.supplier = supplier or existing.supplier
                updated += 1
            else:
                product = Product(
                    store_id=store_id,
                    sku=sku,
                    name=sku,
                    category=category,
                    price_sale=price_sale,
                    price_cost=price_cost,
                    supplier=supplier,
                    units_per_box=units_per_box if units_per_box else 1,
                )
                db.add(product)
                inserted += 1
        except Exception:
            errors += 1

    db.commit()
    return {"total_rows": len(df), "inserted": inserted, "updated": updated, "errors": errors}


def parse_combos_excel(content: bytes, store_id: str, db: Session) -> dict:
    """Parse Listado de combos tiktok.xlsx. Creates combo + items."""
    import pandas as pd
    df = pd.read_excel(io.BytesIO(content))
    df.columns = df.columns.str.strip()

    sku_col = 'SKU SELLER' if 'SKU SELLER' in df.columns else 'SKUID'
    product_cols = [c for c in df.columns if c.startswith('Product')]

    inserted, updated, errors = 0, 0, 0

    # Build product SKU -> ID map for this store
    products = db.query(Product).filter(Product.store_id == store_id).all()
    product_map = {p.sku.lower(): p.id for p in products}
    for p in products:
        product_map[p.name.lower()] = p.id

    for _, row in df.iterrows():
        try:
            combo_sku = _safe_str(row.get(sku_col))
            if not combo_sku:
                errors += 1
                continue

            components = []
            for pcol in product_cols:
                val = _safe_str(row.get(pcol))
                if val:
                    components.append(val)

            if not components:
                errors += 1
                continue

            existing = db.query(Combo).filter(and_(
                Combo.store_id == store_id,
                Combo.combo_sku == combo_sku,
            )).first()

            combo_name = _safe_str(row.get('Nombre combo', row.get('NOMBRE', ''))) or combo_sku

            # Count repeated products correctly: Product1=A, Product2=A → quantity=2 for A
            # This matches V1 data_model.py behavior where each ProductN column = 1 unit
            component_qty: dict[str, int] = {}
            for comp in components:
                pid = product_map.get(comp.lower())
                if pid:
                    component_qty[pid] = component_qty.get(pid, 0) + 1

            if existing:
                existing.combo_name = combo_name
                db.query(ComboItem).filter(ComboItem.combo_id == existing.id).delete()
                db.flush()
                for pid, qty in component_qty.items():
                    db.add(ComboItem(combo_id=existing.id, product_id=pid, quantity=qty))
                updated += 1
            else:
                combo = Combo(store_id=store_id, combo_sku=combo_sku, combo_name=combo_name)
                db.add(combo)
                db.flush()
                for pid, qty in component_qty.items():
                    db.add(ComboItem(combo_id=combo.id, product_id=pid, quantity=qty))
                inserted += 1
        except Exception:
            errors += 1

    db.commit()
    return {"total_rows": len(df), "inserted": inserted, "updated": updated, "errors": errors}


def parse_initial_inventory_excel(content: bytes, store_id: str, db: Session) -> dict:
    """Parse Inventario inicial.xlsx.

    - Tienda SIN inventario inicial: importa y crea los registros (UPSERT por product_id).
    - Tienda CON inventario inicial ya cargado: bloquea el import y devuelve un warning.
      Para ajustar stocks usar el panel de Gestión → Inventario, no reimportar.
    - Nunca borra registros existentes.
    """
    import pandas as pd

    # Block if store already has initial inventory loaded
    existing_count = db.query(InitialInventory).filter(InitialInventory.store_id == store_id).count()
    if existing_count > 0:
        return {
            "total_rows": 0,
            "inserted": 0,
            "updated": 0,
            "errors": 0,
            "unknown_skus": [],
            "warning": (
                f"Esta tienda ya tiene {existing_count} registros de inventario inicial cargados. "
                "Este import solo está disponible para tiendas nuevas. "
                "Para modificar cantidades, usa el panel de Gestión → Inventario Inicial."
            ),
        }

    df = pd.read_excel(io.BytesIO(content))
    df.columns = df.columns.str.strip()

    products = db.query(Product).filter(Product.store_id == store_id).all()
    product_map = {p.sku.lower(): p.id for p in products}
    for p in products:
        product_map[p.name.lower()] = p.id

    inserted, updated, errors = 0, 0, 0
    unknown_skus: list[str] = []

    # First pass: detect unknown SKUs before touching the DB
    for _, row in df.iterrows():
        producto = _safe_str(row.get('Producto') or row.get('ProductoNombre'))
        if not producto:
            errors += 1
            continue
        if not product_map.get(producto.lower()):
            unknown_skus.append(producto)

    if unknown_skus:
        return {
            "total_rows": len(df),
            "inserted": 0,
            "updated": 0,
            "errors": len(unknown_skus),
            "unknown_skus": unknown_skus,
        }

    from datetime import date as date_type
    start_date = date_type(2026, 1, 1)

    # UPSERT: update quantity if product already has a record, else insert
    for _, row in df.iterrows():
        try:
            producto = _safe_str(row.get('Producto') or row.get('ProductoNombre'))
            if not producto:
                errors += 1
                continue
            product_id = product_map.get(producto.lower())
            quantity = _safe_int(row.get('Initial_Stock', row.get('total', row.get('Total', row.get('Cantidad', 0)))))

            existing = db.query(InitialInventory).filter(
                InitialInventory.store_id == store_id,
                InitialInventory.product_id == product_id,
            ).first()

            if existing:
                existing.quantity = quantity
                updated += 1
            else:
                db.add(InitialInventory(
                    store_id=store_id, product_id=product_id,
                    quantity=quantity, start_date=start_date,
                ))
                inserted += 1
        except Exception:
            errors += 1

    db.commit()
    return {"total_rows": len(df), "inserted": inserted, "updated": updated, "errors": errors, "unknown_skus": []}


def parse_amazon_txt(content: bytes, store_id: str, db: Session, batch_id: str | None = None) -> dict:
    """Parse Amazon order report TXT (tab-separated). UPSERTS Amazon orders — never deletes history.
    Multi-brand aware: resuelve brand_id por SKU o prefijo."""
    import uuid as _uuid
    import pandas as pd
    from sqlalchemy import text as _text
    from sqlalchemy.dialects.postgresql import insert as pg_insert

    # Brand map refresh
    _BRAND_MAP_CACHE.pop(store_id, None)
    _sku_map_brands, _prefix_map = _get_brand_maps(db, store_id)

    # Load SKU map for this store
    sku_rows = db.execute(_text("""
        SELECT m.amazon_sku, p.name AS product_name, m.units_per_sale
        FROM amazon_sku_map m
        LEFT JOIN products p ON p.id = m.product_id
        WHERE m.store_id = :sid
    """), {"sid": store_id}).fetchall()
    sku_map = {
        r.amazon_sku: {"product_name": r.product_name, "units_per_sale": r.units_per_sale or 1}
        for r in sku_rows
    }

    df = pd.read_csv(
        io.BytesIO(content), sep='\t', encoding='utf-8-sig',
        on_bad_lines='skip', engine='python'
    )
    df.columns = df.columns.str.strip()

    rows = []
    errors = 0

    for _, row in df.iterrows():
        try:
            order_id = _safe_str(row.get('amazon-order-id'))
            if not order_id:
                errors += 1
                continue

            item_status = _safe_str(row.get('item-status')) or ''
            order_status_raw = _safe_str(row.get('order-status')) or ''

            if 'Cancel' in item_status or 'Cancel' in order_status_raw:
                status = 'Cancelled'
            elif item_status == 'Shipped':
                status = 'Shipped'
            elif item_status == 'Unshipped':
                status = 'Awaiting Shipment'
            else:
                status = item_status or order_status_raw or 'Pending'

            amazon_sku = _safe_str(row.get('sku')) or ''
            qty_ordered = _safe_int(row.get('quantity', 1))

            mapping = sku_map.get(amazon_sku, {})
            mapped_product_name = mapping.get('product_name')
            units_per_sale = mapping.get('units_per_sale', 1)
            expanded_qty = qty_ordered * units_per_sale

            item_price = _safe_float(row.get('item-price')) or 0.0
            shipping_price = _safe_float(row.get('shipping-price')) or 0.0
            promo_discount = _safe_float(row.get('item-promotion-discount')) or 0.0
            purchase_date = _safe_datetime(row.get('purchase-date'))

            rows.append(dict(
                id=str(_uuid.uuid4()),
                store_id=store_id,
                tiktok_order_id=order_id,
                order_date=purchase_date,
                sku=_safe_str(row.get('asin')),
                seller_sku=amazon_sku,
                product_name=mapped_product_name or _safe_str(row.get('product-name')),
                quantity=expanded_qty,
                status=status,
                substatus=item_status,
                price=item_price / max(qty_ordered, 1),
                shipped_time=None,
                created_time=purchase_date,
                sku_subtotal_after_discount=item_price,
                order_amount=item_price + shipping_price,
                order_refund_amount=0.0,
                shipping_fee_after_discount=shipping_price,
                original_shipping_fee=shipping_price,
                sku_seller_discount=0.0,
                sku_platform_discount=promo_discount,
                cancelation_return_type=None,
                fulfillment_type='Merchant',
                buyer_username=None,
                variation=None,
                recipient=None,
                city=_safe_str(row.get('ship-city')),
                state=_safe_str(row.get('ship-state')),
                platform='amazon',
                import_batch_id=batch_id,
                raw_data=None,
                brand_id=_resolve_brand_id(amazon_sku, _sku_map_brands, _prefix_map),
            ))
        except Exception:
            errors += 1

    # UPSERT by (store_id, tiktok_order_id, sku) — never deletes historical Amazon orders.
    # Uploading a partial file (e.g. last month) only adds/updates those rows; history stays intact.
    _update_cols = [
        'product_name', 'quantity', 'status', 'substatus', 'price',
        'sku_subtotal_after_discount', 'order_amount', 'order_refund_amount',
        'shipping_fee_after_discount', 'original_shipping_fee', 'sku_platform_discount',
        'city', 'state', 'import_batch_id', 'brand_id',
    ]

    BATCH = 500   # micro-batches: evita statement_timeout Supabase Micro
    total_processed = 0
    for i in range(0, len(rows), BATCH):
        batch = rows[i:i + BATCH]
        stmt = pg_insert(SalesOrder).values(batch)
        stmt = stmt.on_conflict_do_update(
            constraint='uq_store_order_sku',
            set_={col: getattr(stmt.excluded, col) for col in _update_cols},
        )
        db.execute(stmt)
        db.flush()
        total_processed += len(batch)

    db.commit()
    return {"total_rows": len(df), "inserted": total_processed, "updated": 0, "errors": errors}


def parse_pending_inventory_excel(content: bytes, store_id: str, db: Session) -> dict:
    """Parse Inventario pendiente de recibir.xlsx. APPENDS new records — never deletes existing ones.
    Blocks import if any SKU is not found in the product catalog.
    """
    import pandas as pd
    df = pd.read_excel(io.BytesIO(content))
    df.columns = df.columns.str.strip()

    products = db.query(Product).filter(Product.store_id == store_id).all()
    product_map = {p.sku.lower(): p.id for p in products}
    for p in products:
        product_map[p.name.lower()] = p.id

    inserted, errors = 0, 0
    unknown_skus: list[str] = []

    # First pass: detect unknown SKUs before touching the DB
    for _, row in df.iterrows():
        producto = _safe_str(row.get('Producto'))
        if not producto:
            errors += 1
            continue
        if not product_map.get(producto.lower()):
            unknown_skus.append(producto)

    if unknown_skus:
        return {
            "total_rows": len(df),
            "inserted": 0,
            "updated": 0,
            "errors": len(unknown_skus),
            "unknown_skus": unknown_skus,
        }

    for _, row in df.iterrows():
        try:
            producto = _safe_str(row.get('Producto'))
            if not producto:
                errors += 1
                continue

            product_id = product_map.get(producto.lower())
            qty = _safe_int(row.get('Unidades pedidas', row.get('Cantidad', 0)))
            status_val = _safe_str(row.get('Status', row.get('Estado', 'pending'))) or 'pending'
            supplier = _safe_str(row.get('Proveedor'))
            tracking = _safe_str(row.get('Tracking'))
            cost = _safe_float(row.get('Coste', row.get('Precio')))
            notes = _safe_str(row.get('Notas', row.get('Notes')))

            order_date_val = row.get('Fecha pedido', row.get('Fecha', None))
            order_date = None
            if order_date_val is not None and pd.notna(order_date_val):
                try:
                    dt = pd.to_datetime(order_date_val)
                    if pd.notna(dt):
                        order_date = dt.date()
                except Exception:
                    pass

            db.add(IncomingStock(
                store_id=store_id,
                product_id=product_id,
                qty_ordered=qty,
                order_date=order_date,
                status=status_val,
                supplier=supplier,
                tracking=tracking,
                cost=cost,
                notes=notes,
            ))
            inserted += 1
        except Exception as exc:
            logger.exception("IncomingStock row error — producto=%s exc=%s", producto, exc)
            errors += 1

    db.commit()
    return {"total_rows": len(df), "inserted": inserted, "updated": 0, "errors": errors, "unknown_skus": []}


def parse_walmart_xlsx(content: bytes, store_id: str, db: Session, batch_id: str | None = None) -> dict:
    """Parse Walmart Seller PO Data export (.xlsx). UPSERTS by Order# + SKU — never deletes history.
    Multi-brand aware: resuelve brand_id por SKU o prefijo."""
    import uuid as _uuid
    import openpyxl
    import io as _io
    from sqlalchemy import text as _text
    from sqlalchemy.dialects.postgresql import insert as pg_insert
    from datetime import datetime

    # Brand map fresh
    _BRAND_MAP_CACHE.pop(store_id, None)
    _sku_map_brands, _prefix_map = _get_brand_maps(db, store_id)

    # Load Walmart SKU map
    sku_rows = db.execute(_text("""
        SELECT m.walmart_sku, p.name AS product_name, m.units_per_sale
        FROM walmart_sku_map m
        LEFT JOIN products p ON p.id = m.product_id
        WHERE m.store_id = :sid
    """), {"sid": store_id}).fetchall()
    sku_map = {
        r.walmart_sku: {"product_name": r.product_name, "units_per_sale": r.units_per_sale or 1}
        for r in sku_rows
    }

    wb = openpyxl.load_workbook(_io.BytesIO(content), data_only=True, read_only=True)
    # La hoja se llama 'Po Details'; si no existe, usa la primera
    sheet_name = "Po Details" if "Po Details" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    # Map column name → index (row 1 = headers)
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_idx = {h: i for i, h in enumerate(header_row) if h}

    def _get(row_vals, col_name):
        i = col_idx.get(col_name)
        return row_vals[i] if i is not None and i < len(row_vals) else None

    rows = []
    errors = 0
    for row_vals in ws.iter_rows(min_row=2, values_only=True):
        if not row_vals or not any(row_vals):
            continue
        try:
            order_num = _safe_str(_get(row_vals, "Order#"))
            if not order_num:
                errors += 1
                continue

            walmart_sku = _safe_str(_get(row_vals, "SKU")) or ""
            qty_ordered = _safe_int(_get(row_vals, "Qty")) or 1

            mapping = sku_map.get(walmart_sku, {})
            mapped_name = mapping.get("product_name")
            units_per_sale = mapping.get("units_per_sale", 1)
            expanded_qty = qty_ordered * units_per_sale

            wm_status = _safe_str(_get(row_vals, "Status")) or ""
            if "Cancel" in wm_status:
                status = "Cancelled"
            elif wm_status in ("Shipped", "Delivered"):
                status = wm_status
            elif wm_status:
                status = wm_status
            else:
                status = "Pending"

            order_date_raw = _get(row_vals, "Order Date")
            if isinstance(order_date_raw, datetime):
                order_date = order_date_raw
            else:
                order_date = _safe_datetime(order_date_raw)

            item_cost = _safe_float(_get(row_vals, "Item Cost")) or 0.0
            shipping_cost = _safe_float(_get(row_vals, "Shipping Cost")) or 0.0
            discount = _safe_float(_get(row_vals, "Discount")) or 0.0
            wm_funded = _safe_float(_get(row_vals, "Walmart Funded Incentive")) or 0.0
            tax_val = _safe_float(_get(row_vals, "Tax")) or 0.0

            fulfillment = _safe_str(_get(row_vals, "Fulfillment Entity")) or "SellerFulfilled"

            rows.append(dict(
                id=str(_uuid.uuid4()),
                store_id=store_id,
                tiktok_order_id=order_num,           # reuso del campo como Order#
                order_date=order_date,
                sku=_safe_str(_get(row_vals, "Item ID")) or _safe_str(_get(row_vals, "UPC")),
                seller_sku=walmart_sku,
                product_name=mapped_name or _safe_str(_get(row_vals, "Item Description")),
                quantity=expanded_qty,
                status=status,
                substatus=_safe_str(_get(row_vals, "Service Status")),
                price=item_cost / max(qty_ordered, 1),
                shipped_time=None,
                created_time=order_date,
                sku_subtotal_after_discount=item_cost,
                order_amount=item_cost + shipping_cost + tax_val,
                order_refund_amount=0.0,
                shipping_fee_after_discount=shipping_cost,
                original_shipping_fee=shipping_cost,
                sku_seller_discount=discount,
                sku_platform_discount=wm_funded,
                cancelation_return_type=None,
                fulfillment_type=fulfillment,        # WFSFulfilled | SellerFulfilled
                buyer_username=_safe_str(_get(row_vals, "Customer Name")),
                variation=None,
                recipient=_safe_str(_get(row_vals, "Customer Name")),
                city=_safe_str(_get(row_vals, "City")),
                state=_safe_str(_get(row_vals, "State")),
                platform="walmart",
                import_batch_id=batch_id,
                raw_data=None,
                brand_id=_resolve_brand_id(walmart_sku, _sku_map_brands, _prefix_map),
            ))
        except Exception:
            errors += 1

    if not rows:
        wb.close()
        return {"total_rows": 0, "inserted": 0, "updated": 0, "errors": errors}

    _update_cols = [
        "product_name", "quantity", "status", "substatus", "price",
        "sku_subtotal_after_discount", "order_amount",
        "shipping_fee_after_discount", "original_shipping_fee",
        "sku_seller_discount", "sku_platform_discount",
        "fulfillment_type", "buyer_username", "recipient", "city", "state",
        "import_batch_id",
    ]

    BATCH = 3000
    total_processed = 0
    for i in range(0, len(rows), BATCH):
        batch = rows[i:i + BATCH]
        stmt = pg_insert(SalesOrder).values(batch)
        stmt = stmt.on_conflict_do_update(
            constraint="uq_store_order_sku",
            set_={col: getattr(stmt.excluded, col) for col in _update_cols},
        )
        db.execute(stmt)
        db.flush()
        total_processed += len(batch)

    db.commit()
    wb.close()
    return {"total_rows": len(rows), "inserted": total_processed, "updated": 0, "errors": errors}



# ═════════════════════════════════════════════════════════════════════════════
#  TikTok MERCHANT STATEMENT (Profit & Loss) — settlement + payments
#  Fuente: TikTok Seller Center → Finance → Merchant Statement → P&L
#  UPSERT por (store_id, order_id, sku_id) → subir "last month" no duplica
# ═════════════════════════════════════════════════════════════════════════════
def parse_tiktok_statement_xlsx(content: bytes, store_id: str, db: Session, batch_id: str | None = None) -> dict:
    """Parse TikTok Merchant Statement XLSX (2 sheets: Orders + Order payment info).
    - JOIN por order_id con sales_orders → resuelve brand_id automáticamente
    - UPSERT en tiktok_statement_lines + agrega cabecera en tiktok_statements
    - Idempotente: re-subir mismo mes solo actualiza status/settled_date"""
    import uuid as _uuid
    import openpyxl as _xl
    import io as _io
    from datetime import datetime, date
    from sqlalchemy import text as _text
    from sqlalchemy.dialects.postgresql import insert as pg_insert
    from app.models.tiktok_statement import TiktokStatement, TiktokStatementLine

    def _to_date(v):
        if not v: return None
        try:
            s = str(int(v))  # 20260718 → "20260718"
            if len(s) == 8:
                return date(int(s[:4]), int(s[4:6]), int(s[6:8]))
        except Exception:
            pass
        if isinstance(v, datetime): return v.date()
        if isinstance(v, date): return v
        return None

    def _num(v):
        try: return float(v) if v not in (None, "") else 0.0
        except Exception: return 0.0

    wb = _xl.load_workbook(_io.BytesIO(content), data_only=True)
    if "Orders" not in wb.sheetnames or "Order payment info" not in wb.sheetnames:
        raise ValueError("Excel must have 'Orders' and 'Order payment info' sheets")

    # Sheet 1: Orders (headers row 6, data row 7+)
    ws1 = wb["Orders"]
    h1 = [c for c in next(ws1.iter_rows(min_row=6, max_row=6, values_only=True))]
    idx = {name: i for i, name in enumerate(h1) if name}
    def g1(row, name): return row[idx[name]] if name in idx else None

    # Sheet 2: Order payment info (headers row 1, data row 2+)
    ws2 = wb["Order payment info"]
    h2 = [c for c in next(ws2.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx2 = {name: i for i, name in enumerate(h2) if name}
    def g2(row, name): return row[idx2[name]] if name in idx2 else None

    # Build payment info map keyed by (order_id, sku_id)
    pay_map = {}
    for row in ws2.iter_rows(min_row=2, values_only=True):
        oid = row[idx2["Order ID"]] if "Order ID" in idx2 else None
        sid = row[idx2["SKU ID"]] if "SKU ID" in idx2 else None
        if not oid: continue
        pay_map[(str(oid), str(sid) if sid else None)] = row

    # Cargar brand map (SKU→brand_id) para resolver via sales_orders JOIN
    from sqlalchemy import text
    brand_by_order = {}
    order_ids = list(set(str(r[idx["Order ID"]]) for r in ws1.iter_rows(min_row=7, values_only=True) if r[idx["Order ID"]]))
    if order_ids:
        BATCH_SIZE = 500
        for i in range(0, len(order_ids), BATCH_SIZE):
            batch = order_ids[i:i+BATCH_SIZE]
            placeholders = ",".join([f":o{j}" for j in range(len(batch))])
            params = {"sid": store_id, **{f"o{j}": v for j, v in enumerate(batch)}}
            rows = db.execute(text(f"""
                SELECT DISTINCT tiktok_order_id, brand_id
                FROM sales_orders
                WHERE store_id = :sid AND tiktok_order_id IN ({placeholders})
                  AND brand_id IS NOT NULL
            """), params).fetchall()
            for tk_oid, bid in rows:
                brand_by_order[str(tk_oid)] = bid

    # Build lines
    lines = []
    stmt_agg = {}  # statement_id → aggregates
    for row in ws1.iter_rows(min_row=7, values_only=True):
        oid = row[idx["Order ID"]] if "Order ID" in idx else None
        if not oid: continue
        oid = str(oid)
        sid = str(row[idx["SKU ID"]]) if "SKU ID" in idx and row[idx["SKU ID"]] else None
        stmt = str(row[idx["linked statement id"]]) if "linked statement id" in idx and row[idx["linked statement id"]] else None
        payout = str(row[idx["linked payout id"]]) if "linked payout id" in idx and row[idx["linked payout id"]] else None

        pay = pay_map.get((oid, sid), [None]*len(h2))

        income = _num(g1(row, "Order Income"))
        cost = _num(g1(row, "Order Cost"))
        margin = _num(g1(row, "Net Order Margin"))
        # Fees (guardamos como positivos, TikTok los da negativos)
        referral = abs(_num(g1(row, "Referral fee")))
        smart_promo = abs(_num(g1(row, "Smart Promotion fee")))
        smart_camp = abs(_num(g1(row, "Campaign resource fee")))
        managed_svc = abs(_num(g1(row, "Managed service plan (Per order fee)")))
        tt_ship = abs(_num(g1(row, "TikTok Shop shipping fee")))
        fbt_ship = abs(_num(g1(row, "Fulfilled by TikTok Shop shipping fee")))
        tt_incentive = abs(_num(g1(row, "TikTok Shop shipping incentive")))
        aff_comm = abs(_num(g1(row, "Affiliate Commission"))) + abs(_num(g1(row, "Affiliate partner commission")))

        line = dict(
            id=str(_uuid.uuid4()),
            store_id=store_id,
            statement_id=stmt,
            payout_id=payout,
            order_id=oid,
            sku_id=sid,
            product_name=str(g1(row, "Product name") or "")[:500],
            order_income=income,
            order_cost=cost,
            net_order_margin=margin,
            sold_quantity=int(g1(row, "Sold Quantity") or 0),
            order_paid_date=_to_date(g1(row, "Order paid date")),
            order_shipment_date=_to_date(g1(row, "Order shipment date")),
            order_delivery_date=_to_date(g1(row, "Order delivery date")),
            order_settled_date=_to_date(g1(row, "Order settled date")),
            order_status=str(g1(row, "Order status") or "")[:30],
            unsettled_reason=str(g1(row, "unsettled reasons") or "")[:200],
            gross_sales=_num(g1(row, "Gross sales")),
            seller_discount=_num(g1(row, "Seller discount")),
            gross_sales_refund=_num(g1(row, "Gross sales refund")),
            referral_fee=referral,
            smart_promo_fee=smart_promo,
            smart_promo_camp_fee=smart_camp,
            managed_service_fee=managed_svc,
            tiktok_shipping_fee=tt_ship,
            fbt_shipping_fee=fbt_ship,
            tiktok_ship_incentive=tt_incentive,
            affiliate_commission=aff_comm,
            customer_paid_ship=_num(g1(row, "Customer-paid shipping fee")),
            sku_subtotal_before=_num(g2(pay, "SKU Subtotal Before Discount")),
            sku_subtotal_after=_num(g2(pay, "SKU Subtotal After Discount")),
            order_amount=_num(g2(pay, "Order Amount")),
            taxes=_num(g2(pay, "Taxes")),
            brand_id=brand_by_order.get(oid),
            import_batch_id=batch_id,
        )
        lines.append(line)

        # Agregar por statement
        if stmt:
            agg = stmt_agg.setdefault(stmt, {
                "statement_id": stmt, "payout_id": payout,
                "total_income": 0, "total_cost": 0, "total_margin": 0,
                "total_fees": 0, "total_orders": set(),
                "period_start": None, "period_end": None, "settled_date": None,
            })
            agg["total_income"] += income
            agg["total_cost"] += cost
            agg["total_margin"] += margin
            agg["total_fees"] += referral + smart_promo + smart_camp + managed_svc + fbt_ship + tt_ship
            agg["total_orders"].add(oid)
            pd = line["order_paid_date"]
            sd = line["order_settled_date"]
            if pd:
                if agg["period_start"] is None or pd < agg["period_start"]: agg["period_start"] = pd
                if agg["period_end"] is None or pd > agg["period_end"]: agg["period_end"] = pd
            if sd:
                if agg["settled_date"] is None or sd > agg["settled_date"]: agg["settled_date"] = sd

    # UPSERT lines
    line_update_cols = [
        "statement_id", "payout_id", "product_name", "order_income", "order_cost",
        "net_order_margin", "sold_quantity", "order_paid_date", "order_shipment_date",
        "order_delivery_date", "order_settled_date", "order_status", "unsettled_reason",
        "gross_sales", "seller_discount", "gross_sales_refund", "referral_fee",
        "smart_promo_fee", "smart_promo_camp_fee", "managed_service_fee",
        "tiktok_shipping_fee", "fbt_shipping_fee", "tiktok_ship_incentive",
        "affiliate_commission", "customer_paid_ship", "sku_subtotal_before",
        "sku_subtotal_after", "order_amount", "taxes", "brand_id", "import_batch_id",
    ]
    # Deduplicar por (store_id, order_id, sku_id) — mantener último — antes del UPSERT
    # Postgres ON CONFLICT no soporta múltiples rows con misma target key en un batch.
    _dedup = {}
    for L in lines:
        _key = (L.get("store_id"), L.get("order_id"), L.get("sku_id"))
        _dedup[_key] = L  # último gana
    lines_deduped = list(_dedup.values())

    BATCH = 500
    upserted_lines = 0
    for i in range(0, len(lines_deduped), BATCH):
        batch = lines_deduped[i:i+BATCH]
        stmt_sql = pg_insert(TiktokStatementLine).values(batch)
        stmt_sql = stmt_sql.on_conflict_do_update(
            constraint="uq_tt_line",
            set_={col: getattr(stmt_sql.excluded, col) for col in line_update_cols},
        )
        db.execute(stmt_sql)
        db.flush()
        upserted_lines += len(batch)

    # UPSERT statement headers
    stmt_headers = []
    for sid, agg in stmt_agg.items():
        stmt_headers.append(dict(
            id=str(_uuid.uuid4()),
            store_id=store_id,
            statement_id=sid,
            payout_id=agg["payout_id"],
            total_income=agg["total_income"],
            total_cost=agg["total_cost"],
            total_margin=agg["total_margin"],
            total_fees=agg["total_fees"],
            total_orders=len(agg["total_orders"]),
            period_start=agg["period_start"],
            period_end=agg["period_end"],
            settled_date=agg["settled_date"],
            import_batch_id=batch_id,
        ))
    hdr_update_cols = ["payout_id", "total_income", "total_cost", "total_margin",
                       "total_fees", "total_orders", "period_start", "period_end",
                       "settled_date", "import_batch_id"]
    upserted_stmts = 0
    for i in range(0, len(stmt_headers), 100):
        batch = stmt_headers[i:i+100]
        if not batch: continue
        stmt_sql = pg_insert(TiktokStatement).values(batch)
        stmt_sql = stmt_sql.on_conflict_do_update(
            constraint="uq_tt_stmt",
            set_={col: getattr(stmt_sql.excluded, col) for col in hdr_update_cols},
        )
        db.execute(stmt_sql)
        db.flush()
        upserted_stmts += len(batch)

    db.commit()
    matched = sum(1 for l in lines if l["brand_id"])
    return {
        "total_lines": len(lines),
        "upserted_lines": upserted_lines,
        "statements": upserted_stmts,
        "matched_to_orders": matched,
        "unmatched": len(lines) - matched,
    }
