import io
import logging
import uuid
from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import desc

from app.config import MAX_UPLOAD_SIZE
from app.database import get_db, SessionLocal
from app.models.user import User
from app.models.import_history import ImportHistory
from app.dependencies import get_current_user
from app.schemas.import_schemas import ImportResult
from app.services.import_service import (
    parse_orders_csv,
    parse_affiliate_csv,
    parse_products_excel,
    parse_combos_excel,
    parse_initial_inventory_excel,
    parse_pending_inventory_excel,
    parse_amazon_txt,
)

logger = logging.getLogger("rodmat.imports")
router = APIRouter(prefix="/api/import", tags=["import"])


def _validate_upload(file: UploadFile, content: bytes, allowed_ext: str):
    if len(content) > MAX_UPLOAD_SIZE:
        raise HTTPException(status_code=413, detail=f"File too large. Maximum size is {MAX_UPLOAD_SIZE // (1024*1024)}MB")
    filename = (file.filename or "").lower()
    if not filename.endswith(allowed_ext):
        raise HTTPException(status_code=415, detail=f"Invalid file type. Expected {allowed_ext}")


def _target_store(user: User, store_id: str | None) -> str:
    """Superadmin puede pasar store_id explícito; cualquier otro usa su propia tienda."""
    if user.role == "superadmin" and store_id:
        return store_id
    return user.store_id


@router.post("/orders", response_model=ImportResult)
async def import_orders(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    store_id: str | None = Query(None),
    send_report: bool = Query(False, description="Send daily report after import"),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from app.services.analytics_service import _cache
    content = await file.read()
    _validate_upload(file, content, ".csv")
    target = _target_store(user, store_id)

    batch_id = str(uuid.uuid4())
    history = ImportHistory(id=batch_id, store_id=target, import_type="tiktok",
                            filename=file.filename, imported_by=user.email)
    db.add(history)
    db.flush()

    result = parse_orders_csv(content, target, db, batch_id=batch_id)

    history.rows_imported = result["inserted"]
    history.rows_deleted = result.get("rows_deleted", 0)
    db.commit()

    _cache.clear()
    if send_report:
        from app.api.reports import _send_report_bg
        background_tasks.add_task(_send_report_bg, target)
        logger.info("Post-import report queued for store %s", target[:8])
    return result


@router.post("/affiliates", response_model=ImportResult)
async def import_affiliates(
    file: UploadFile = File(...),
    store_id: str | None = Query(None),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    content = await file.read()
    _validate_upload(file, content, ".csv")
    return parse_affiliate_csv(content, _target_store(user, store_id), db)


@router.post("/products", response_model=ImportResult)
async def import_products(
    file: UploadFile = File(...),
    store_id: str | None = Query(None),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    content = await file.read()
    _validate_upload(file, content, ".xlsx")
    return parse_products_excel(content, _target_store(user, store_id), db)


@router.post("/combos", response_model=ImportResult)
async def import_combos(
    file: UploadFile = File(...),
    store_id: str | None = Query(None),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    content = await file.read()
    _validate_upload(file, content, ".xlsx")
    return parse_combos_excel(content, _target_store(user, store_id), db)


@router.post("/initial-inventory", response_model=ImportResult)
async def import_initial_inventory(
    file: UploadFile = File(...),
    store_id: str | None = Query(None),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    content = await file.read()
    _validate_upload(file, content, ".xlsx")
    return parse_initial_inventory_excel(content, _target_store(user, store_id), db)


@router.post("/incoming-stock", response_model=ImportResult)
async def import_incoming_stock(
    file: UploadFile = File(...),
    store_id: str | None = Query(None),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    content = await file.read()
    _validate_upload(file, content, ".xlsx")
    return parse_pending_inventory_excel(content, _target_store(user, store_id), db)


@router.post("/amazon", response_model=ImportResult)
async def import_amazon_orders(
    file: UploadFile = File(...),
    store_id: str | None = Query(None),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from app.services.analytics_service import _cache, _df_cache
    content = await file.read()
    if len(content) > MAX_UPLOAD_SIZE:
        raise HTTPException(status_code=413, detail="File too large.")
    filename = (file.filename or "").lower()
    if not (filename.endswith(".txt") or filename.endswith(".tsv") or filename.endswith(".csv")):
        raise HTTPException(status_code=415, detail="Expected .txt or .tsv file from Amazon Seller Central.")

    target = _target_store(user, store_id)
    batch_id = str(uuid.uuid4())
    history = ImportHistory(id=batch_id, store_id=target, import_type="amazon",
                            filename=file.filename, imported_by=user.email)
    db.add(history)
    db.flush()

    result = parse_amazon_txt(content, target, db, batch_id=batch_id)

    history.rows_imported = result["inserted"]
    history.rows_deleted = result.get("rows_deleted", 0)
    db.commit()

    _cache.clear()
    _df_cache.clear()
    return result


@router.get("/history")
async def get_import_history(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    limit: int = Query(50, ge=1, le=200),
):
    q = db.query(ImportHistory)
    if user.role != "superadmin":
        q = q.filter(ImportHistory.store_id == user.store_id)
    rows = q.order_by(desc(ImportHistory.imported_at)).limit(limit).all()
    return [
        {
            "id": r.id,
            "import_type": r.import_type,
            "filename": r.filename,
            "rows_imported": r.rows_imported,
            "rows_deleted": r.rows_deleted,
            "imported_by": r.imported_by,
            "imported_at": r.imported_at.isoformat() if r.imported_at else None,
        }
        for r in rows
    ]


@router.delete("/history/{batch_id}")
async def delete_import_batch(
    batch_id: str,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from sqlalchemy import text as _text

    history = db.query(ImportHistory).filter(ImportHistory.id == batch_id).first()
    if not history:
        raise HTTPException(status_code=404, detail="Import batch not found")
    if user.role != "superadmin" and history.store_id != user.store_id:
        raise HTTPException(status_code=403, detail="Access denied")
    if history.import_type not in ("tiktok", "amazon"):
        raise HTTPException(status_code=400, detail=f"Cannot rollback import type '{history.import_type}'")

    del_result = db.execute(_text("DELETE FROM sales_orders WHERE import_batch_id = :bid"), {"bid": batch_id})
    rows_deleted = del_result.rowcount
    db.delete(history)
    db.commit()

    try:
        from app.services.analytics_service import _cache, _df_cache
        _cache.clear()
        _df_cache.clear()
    except Exception:
        pass

    logger.info("Import batch %s deleted — %d rows removed", batch_id[:8], rows_deleted)
    return {"deleted_rows": rows_deleted, "batch_id": batch_id}


_TEMPLATES: dict[str, dict] = {
    "products": {
        "filename": "plantilla_productos.xlsx",
        "headers": ["Producto", "Coste", "PRECIO", "UNIDADES POR CAJA", "Tipo", "Proveedor"],
        "example": ["PROD-001", 5.50, 12.99, 10, "Electrónica", "Proveedor SA"],
    },
    "combos": {
        "filename": "plantilla_combos.xlsx",
        "headers": ["SKU SELLER", "Nombre combo", "Product1", "Product2", "Product3"],
        "example": ["COMBO-001", "Pack 2 productos", "PROD-001", "PROD-002", ""],
    },
    "initial-inventory": {
        "filename": "plantilla_inventario_inicial.xlsx",
        "headers": ["Producto", "Initial_Stock"],
        "example": ["PROD-001", 100],
    },
    "incoming-stock": {
        "filename": "plantilla_stock_pendiente.xlsx",
        "headers": ["Producto", "Unidades pedidas", "Status", "Proveedor", "Tracking", "Coste", "Notas", "Fecha pedido"],
        "example": ["PROD-001", 50, "pending", "Proveedor SA", "TRACK123", 5.50, "", "2026-05-16"],
    },
}


@router.get("/templates/{template_name}")
async def download_template(template_name: str):
    """Return a blank XLSX template file for the given import type (no auth required)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    if template_name not in _TEMPLATES:
        raise HTTPException(status_code=404, detail=f"Template '{template_name}' not found. Available: {list(_TEMPLATES)}")

    tmpl = _TEMPLATES[template_name]
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla"

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(tmpl["headers"], start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(len(header) + 4, 16)

    for col_idx, value in enumerate(tmpl["example"], start=1):
        ws.cell(row=2, column=col_idx, value=value)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{tmpl["filename"]}"'},
    )
