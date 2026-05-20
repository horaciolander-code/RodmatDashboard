import io
from datetime import datetime, timezone
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.inventory import InitialInventory, IncomingStock, FBTInventory
from app.models.product import Product
from app.models.user import User
from app.schemas.inventory import (
    InitialInventoryCreate, InitialInventoryResponse,
    IncomingStockCreate, IncomingStockResponse,
    IncomingStockUpdate,
    FBTInventoryCreate, FBTInventoryUpdate, FBTInventoryResponse,
)
from app.dependencies import get_current_user

router = APIRouter(prefix="/api/inventory", tags=["inventory"])


@router.post("/initial", response_model=InitialInventoryResponse, status_code=201)
def create_initial_inventory(
    payload: InitialInventoryCreate,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    record = InitialInventory(store_id=user.store_id, **payload.model_dump())
    db.add(record)
    db.commit()
    db.refresh(record)
    return record


@router.get("/initial", response_model=list[InitialInventoryResponse])
def list_initial_inventory(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    return db.query(InitialInventory).filter(InitialInventory.store_id == user.store_id).all()


@router.post("/incoming", response_model=IncomingStockResponse, status_code=201)
def create_incoming_stock(
    payload: IncomingStockCreate,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    record = IncomingStock(store_id=user.store_id, **payload.model_dump())
    db.add(record)
    db.commit()
    db.refresh(record)
    return record


@router.get("/incoming", response_model=list[IncomingStockResponse])
def list_incoming_stock(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    records = db.query(IncomingStock).filter(IncomingStock.store_id == user.store_id).all()
    product_ids = {r.product_id for r in records}
    pmap = {p.id: p.name for p in db.query(Product).filter(Product.id.in_(product_ids)).all()} if product_ids else {}
    result = []
    for r in records:
        d = IncomingStockResponse.model_validate(r)
        d.product_name = pmap.get(r.product_id)
        result.append(d)
    return result


@router.put("/incoming/{record_id}", response_model=IncomingStockResponse)
def update_incoming_stock(
    record_id: str,
    payload: IncomingStockUpdate,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    record = db.query(IncomingStock).filter(
        IncomingStock.id == record_id, IncomingStock.store_id == user.store_id
    ).first()
    if not record:
        raise HTTPException(status_code=404, detail="Incoming stock record not found")
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(record, field, value)
    db.commit()
    db.refresh(record)
    return record


@router.delete("/incoming/{record_id}", status_code=204)
def delete_incoming_stock(
    record_id: str,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    record = db.query(IncomingStock).filter(
        IncomingStock.id == record_id, IncomingStock.store_id == user.store_id
    ).first()
    if not record:
        raise HTTPException(status_code=404, detail="Incoming stock record not found")
    db.delete(record)
    db.commit()


@router.get("/incoming/export")
def export_incoming_stock(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Descarga todos los registros de incoming stock como Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    records = db.query(IncomingStock).filter(IncomingStock.store_id == user.store_id).all()
    product_ids = {r.product_id for r in records}
    pmap = {p.id: p.name for p in db.query(Product).filter(Product.id.in_(product_ids)).all()} if product_ids else {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Pendiente"

    headers = ["Producto", "Unidades pedidas", "Status", "Proveedor", "Tracking", "Coste", "Notas", "Fecha pedido"]
    fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(len(h) + 4, 16)

    for row_idx, r in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=pmap.get(r.product_id) or r.product_id)
        ws.cell(row=row_idx, column=2, value=r.qty_ordered)
        ws.cell(row=row_idx, column=3, value=r.status)
        ws.cell(row=row_idx, column=4, value=r.supplier)
        ws.cell(row=row_idx, column=5, value=r.tracking)
        ws.cell(row=row_idx, column=6, value=r.cost)
        ws.cell(row=row_idx, column=7, value=r.notes)
        ws.cell(row=row_idx, column=8, value=str(r.order_date) if r.order_date else None)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fecha = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="incoming_stock_backup_{fecha}.xlsx"'},
    )


# ── FBT Inventory ─────────────────────────────────────────────────────────────

@router.post("/fbt", response_model=FBTInventoryResponse, status_code=201)
def create_fbt(
    payload: FBTInventoryCreate,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    record = FBTInventory(store_id=user.store_id, **payload.model_dump())
    db.add(record)
    db.commit()
    db.refresh(record)
    return record


@router.get("/fbt", response_model=list[FBTInventoryResponse])
def list_fbt(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    return db.query(FBTInventory).filter(FBTInventory.store_id == user.store_id).all()


@router.put("/fbt/{record_id}", response_model=FBTInventoryResponse)
def update_fbt(
    record_id: str,
    payload: FBTInventoryUpdate,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    record = db.query(FBTInventory).filter(
        FBTInventory.id == record_id, FBTInventory.store_id == user.store_id
    ).first()
    if not record:
        raise HTTPException(status_code=404, detail="FBT record not found")
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(record, field, value)
    db.commit()
    db.refresh(record)
    return record


@router.delete("/fbt/{record_id}", status_code=204)
def delete_fbt(
    record_id: str,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    record = db.query(FBTInventory).filter(
        FBTInventory.id == record_id, FBTInventory.store_id == user.store_id
    ).first()
    if not record:
        raise HTTPException(status_code=404, detail="FBT record not found")
    db.delete(record)
    db.commit()
